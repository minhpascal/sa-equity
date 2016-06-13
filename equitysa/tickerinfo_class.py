import openpyxl
import get_ticker_info 
import filter_bullets
import string 
import glob
import csv
import os


class TickerInfo(object):
    def __init__(self,tag,ticker): #single ticker
        self.ticker = ticker.upper()
        self._tag = tag
        Scraper = get_ticker_info.GetTickerInfo(self.tag,self.ticker)
        self.ma = Scraper() 

    @property 
    def tag(self):
        idioms=  {'news':'/news/on-the-move/',
                  'ma':'/news/m-a/',
                  'dividends': '/news/dividends/',
                  'earnings': '/news/earnings_news/',
                  'general': '/news/',
                  }
        if idioms.get(self._tag,None) is not None:
            return idioms[self._tag]
        else:
            return self._tag

    @classmethod
    def get_attr_at_date(self, datekey_ht, **date):
        '''
        generic class method to extract values from a dictionary 
        with datetime.date objects as keys
        
        year/day/month values of date most be ints
        '''
        res = list()
        for key,value in datekey_ht.items():
            #Python shortcircuits boolean exprs; no KeyError
            is_day,is_month,is_year = [True]*3
            if 'year' in date.keys():
                is_year = date['year'] == key.year
            if 'month' in date.keys():
                is_month = date['month'] == key.month
            if 'day' in date.keys():
                is_day = date['day']== key.day
            if (is_year and is_month and is_day):
                res.append(value)
        return res
    
    def write_to_excel_at_year(self, ws, year, info_ht):
        '''
        info_ht defaults to self.ma
        '''
        def flatten(L, res = None):
            if type(L)!=list:
                return L
            if res==None:
                res =[]
            for i in L:
                if (type(i)!=list):
                    res+=[i]
                else:
                    flatten(i,res)
            return res
        ticker_row = -1
        present_col = -1
        for r in range(1,len(ws.active.rows)+1):
            if (ws.active.cell(row=r,column=1).value == self.ticker):
               ticker_row = r
        for c in range(1,len(ws.active.columns)+1):
            if (ws.active.cell(row=1,column=c).value == year):
                present_col=c
        
        if (present_col==-1): 
            raise Exception("year not found")
        if (ticker_row==-1): 
            raise Exception("Ticker not found")
        L = TickerInfo.get_attr_at_date(info_ht, **{'year':year})
        L = '\n\n'.join(filter_bullets.focus(L))
        if (L==[]):
            L = ''
        ws.active.cell(row=ticker_row, column=present_col).value = L
        


 



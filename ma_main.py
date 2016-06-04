# -*- coding: utf-8 -*-
"""
Created on Fri Jun  3 14:57:44 2016

@author: VineetA
"""
import tickerinfo_class as tick
import openpyxl

def tickerMA_main():
    with open("tickers.txt") as f:
        target_str = f.read()
    tickers = [word.strip() for word in target_str.split(',')]
 
    tk = list()
    for ticker in tickers:
        tk.append(tick.TickerInfo(ticker))
  
    ws = openpyxl.Workbook()
    for i in range(1,8):
        ws.active.cell(row=1, column = i+1).value = 2017-i
    for i,e in enumerate(tickers):
        ws.active.cell(row = i+2, column = 1).value = e
 
    for i,e in enumerate(tk):
        for j in range(2010,2017):
            e.write_to_excel_at_year(ws=ws, year=j,info_ht=e.ma)
    ws.save('C:\\Users\\VineetA\\Documents\\Python Scripts\\equitysa\\sa.xlsx') 


if __name__ == '__main__':
    tickerMA_main()